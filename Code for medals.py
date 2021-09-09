from openpyxl import Workbook,load_workbook 
wb=load_workbook("Medals.xlsx")
ws = wb.active
golds={}
totals={}
silvers={}
bronzes={}
golds_list=[]
silvers_list=[]
bronzes_list=[]
totals_list=[]
for row in range (2,95):
    for column in range (1,8):
       print(str(ws.cell(row,column).value),end="") 
       if column==3:
           golds[str(ws.cell(row,column).value)]=str(ws.cell(row,2).value)
       elif column==4:
           silvers[str(ws.cell(row,column).value)]=str(ws.cell(row,2).value)
       elif column==5:
           bronzes[str(ws.cell(row,column).value)]=str(ws.cell(row,2).value)
       elif column==6:
           totals[str(ws.cell(row,column).value)]=str(ws.cell(row,2).value)
    print()
for key in golds.keys():
    golds_list.append(int(key))
for key in silvers.keys():
    silvers_list.append(int(key))    
for key in bronzes.keys():
    bronzes_list.append(int(key))
for key in totals.keys():
    totals_list.append(int(key))
golds_winner=max(golds_list)
silvers_winner=max(silvers_list)
bronzes_winner=max(bronzes_list)
totals_winner=max(totals_list)
print('The winner in total is',totals[str(totals_winner)],'with',totals_winner )
print('The winner in gold is',golds[str(golds_winner)],'with',golds_winner)
print('The winner in bronze is',bronzes[str(bronzes_winner)],'with',bronzes_winner)
print('The winner in silver is',silvers[str(silvers_winner)],'with',silvers_winner)